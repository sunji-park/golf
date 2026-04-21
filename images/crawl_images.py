"""
골프장 공식 홈페이지에서 이미지를 수집하는 샘플 크롤러.

사용법:
    python crawl_images.py --limit 10 --sheet korea
    python crawl_images.py --sheet japan --delay 3

주의:
    - 각 사이트의 robots.txt 확인 후 사용하세요.
    - 상업적 배포 시 저작권 계약이 필요할 수 있습니다.
    - 요청 간 지연(--delay)을 충분히 두세요.
"""
import argparse
import json
import os
import re
import time
from urllib.parse import urljoin, urlparse

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

HEADERS = {
    "User-Agent": "GolfCourseBot/1.0 (contact@yourdomain.com)"
}

def slugify(name):
    s = re.sub(r'[\\/:*?"<>|]', '', str(name))
    return s.replace(' ', '_').strip('_')[:60]


def read_targets(xlsx_path, sheet_name, name_col, url_col, limit=None):
    wb = load_workbook(xlsx_path, read_only=True)
    ws = wb[sheet_name]
    targets = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        name = row[name_col - 1]
        url = row[url_col - 1]
        if not name or not url or not str(url).startswith("http"):
            continue
        targets.append((str(name), str(url)))
        if limit and len(targets) >= limit:
            break
    wb.close()
    return targets


def extract_image_urls(html, base_url):
    """Extract og:image, <img src>, and hero banner style images."""
    soup = BeautifulSoup(html, "html.parser")
    urls = set()

    # Open Graph image
    og = soup.find("meta", attrs={"property": "og:image"})
    if og and og.get("content"):
        urls.add(urljoin(base_url, og["content"]))

    # Twitter card image
    tw = soup.find("meta", attrs={"name": "twitter:image"})
    if tw and tw.get("content"):
        urls.add(urljoin(base_url, tw["content"]))

    # <img> tags
    for img in soup.find_all("img"):
        src = img.get("src") or img.get("data-src") or img.get("data-original")
        if not src:
            continue
        full = urljoin(base_url, src)
        # Filter obvious icons/thumbs
        if any(k in full.lower() for k in ["icon", "favicon", "pixel", "spacer", "logo_"]):
            continue
        urls.add(full)

    # Inline background-image styles (hero sliders often use these)
    for el in soup.find_all(style=True):
        m = re.search(r"url\(['\"]?(.+?)['\"]?\)", el["style"])
        if m:
            urls.add(urljoin(base_url, m.group(1)))

    return [u for u in urls if u.lower().rsplit(".", 1)[-1].split("?")[0] in
            ("jpg", "jpeg", "png", "webp", "gif")]


def download(url, dest_path, timeout=15):
    try:
        r = requests.get(url, headers=HEADERS, timeout=timeout, stream=True)
        r.raise_for_status()
        if int(r.headers.get("content-length", 0)) > 10 * 1024 * 1024:
            return False  # skip >10MB
        with open(dest_path, "wb") as f:
            for chunk in r.iter_content(8192):
                f.write(chunk)
        return True
    except Exception as e:
        print(f"    [ERR] {url}: {e}")
        return False


def crawl_course(name, homepage, out_dir, delay=2):
    os.makedirs(out_dir, exist_ok=True)
    print(f"\n[+] {name} -> {homepage}")
    try:
        r = requests.get(homepage, headers=HEADERS, timeout=15)
        r.raise_for_status()
    except Exception as e:
        print(f"    [ERR] fetch homepage failed: {e}")
        return 0

    img_urls = extract_image_urls(r.text, homepage)
    print(f"    Found {len(img_urls)} image candidates")

    meta = {"course_name": name, "homepage": homepage, "images": []}
    ok = 0
    for i, url in enumerate(img_urls[:20]):  # cap 20 per course
        ext = url.rsplit(".", 1)[-1].split("?")[0].lower()
        if ext not in ("jpg", "jpeg", "png", "webp"):
            ext = "jpg"
        fname = f"img{i:02d}.{ext}"
        dest = os.path.join(out_dir, fname)
        if download(url, dest):
            meta["images"].append({"filename": fname, "source_url": url})
            ok += 1
        time.sleep(delay)

    with open(os.path.join(out_dir, "_meta.json"), "w", encoding="utf-8") as f:
        json.dump(meta, f, ensure_ascii=False, indent=2)
    print(f"    Downloaded: {ok}/{len(img_urls[:20])}")
    return ok


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", default="../golfcourse_database.xlsx")
    ap.add_argument("--sheet", choices=["korea", "japan"], default="korea")
    ap.add_argument("--limit", type=int, default=10)
    ap.add_argument("--delay", type=float, default=2.0)
    args = ap.parse_args()

    if args.sheet == "korea":
        sheet_name = "코스정보_국내"
        name_col, url_col = 3, 20   # C, T
        out_base = "korea"
    else:
        sheet_name = "코스정보_일본"
        name_col, url_col = 3, 17   # C, Q
        out_base = "japan"

    targets = read_targets(args.xlsx, sheet_name, name_col, url_col, limit=args.limit)
    print(f"Targets: {len(targets)} courses")

    total = 0
    for name, url in targets:
        out_dir = os.path.join(os.path.dirname(__file__), out_base, slugify(name))
        total += crawl_course(name, url, out_dir, delay=args.delay)

    print(f"\nDone. Total images saved: {total}")


if __name__ == "__main__":
    main()
